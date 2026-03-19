"""Chat and messaging routes for conversations, contacts, groups, and messages."""
from collections import defaultdict
import copy
import csv
from io import BytesIO, StringIO
import json
import logging
import os
from datetime import datetime, timedelta
from urllib.parse import urlparse

import requests
from flask import Blueprint, request
from flask_jwt_extended import jwt_required, get_jwt_identity
from sqlalchemy import desc, and_

from app.database import db
from app.models.user import User
from app.models.whatsapp import (
    Contact,
    Conversation,
    Message,
    Template,
    Group,
    GroupContact,
    GroupMessage,
    GroupMessageRecipient,
    WabaAccount,
)
from app.utils.datetime_utils import ist_now
from app.utils.object_storage import upload_file
from app.utils.utils import success_response, error_response

chat_bp = Blueprint('chat', __name__)
logger = logging.getLogger(__name__)

CONVERSATION_WINDOW_HOURS = 24
ALLOWED_IMAGE_MIME_TYPES = {'image/jpeg', 'image/png'}
ALLOWED_IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png'}
MAX_IMAGE_SIZE_BYTES = 5 * 1024 * 1024


def send_wa_text_message(phone_number_id, access_token, payload, is_template=False):
    """Send a WhatsApp message via the Graph API and return the outbound wamid."""
    try:
        base_url = os.getenv('GRAPH_URL_BASE', 'https://graph.facebook.com')
        url_version = os.getenv('GRAPH_URL_VERSION', 'v22.0')
        endpoint = f"{phone_number_id}/messages"
        url = f"{base_url}/{url_version}/{endpoint}"

        headers = {
            'Authorization': f'Bearer {access_token}'
        }

        if is_template:
            payload['template'] = json.dumps(payload.get('template', {}))
        else:
            payload['text'] = json.dumps(payload.get('text', {}))

        response = requests.request('POST', url, headers=headers, data=payload, timeout=20)

        logger.info('******** WA response *************')
        logger.info(response.text)
        logger.info('******** WA response *************')

        if response.status_code not in (200, 201, 202):
            logger.debug(f"Response from WA send text message API: {response.text}")
            return None, response.status_code, response.text

        response_json = response.json()
        wamid = response_json.get('messages', [{}])[0].get('id')
        return wamid, response.status_code, None

    except Exception as e:
        raise ValueError(f"Failed to send whatsapp text message {e}")


# ---------------------------------------------------------------------------
# Request parsing
# ---------------------------------------------------------------------------

def _parse_send_message_request_payload():
    """Parse JSON or multipart/form-data for the send-message endpoint.

    Returns (data_dict, uploaded_file_or_None).
    When the request is plain JSON there is no file upload; the caller must
    check campaign_image_url in data_dict instead.
    """
    if request.is_json:
        return request.get_json() or {}, None

    form_data = request.form.to_dict(flat=True)
    template_raw = form_data.get('template')
    if template_raw is not None:
        try:
            form_data['template'] = json.loads(template_raw)
        except (TypeError, ValueError):
            raise ValueError('template must be valid JSON when using multipart/form-data')

    uploaded_campaign_image = request.files.get('header_image') or request.files.get('file')
    return form_data, uploaded_campaign_image


def _coerce_int(value, field_name):
    if value is None or value == '':
        return None

    if isinstance(value, int):
        return value

    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())

    raise ValueError(f'{field_name} must be an integer')


# ---------------------------------------------------------------------------
# Graph API helpers
# ---------------------------------------------------------------------------


def _validate_campaign_image_file(uploaded_file):
    if not uploaded_file:
        raise ValueError('campaign_image is required for IMAGE header templates')

    filename = (uploaded_file.filename or '').strip()
    extension = os.path.splitext(filename)[1].lower()
    if extension not in ALLOWED_IMAGE_EXTENSIONS:
        raise ValueError('Allowed campaign image types are JPG and PNG only')

    mime_type = (uploaded_file.mimetype or '').lower()
    if mime_type not in ALLOWED_IMAGE_MIME_TYPES:
        raise ValueError('Allowed campaign image MIME types are image/jpeg and image/png only')

    uploaded_file.stream.seek(0, os.SEEK_END)
    file_size = uploaded_file.stream.tell()
    uploaded_file.stream.seek(0)
    if file_size > MAX_IMAGE_SIZE_BYTES:
        raise ValueError('Campaign image size must be 5 MB or smaller')

    return filename, mime_type


def _validate_campaign_image_url(image_url):
    parsed = urlparse(image_url or '')
    if parsed.scheme not in {'http', 'https'} or not parsed.netloc:
        raise ValueError('campaign_image_url must be a valid http/https URL')

    path = (parsed.path or '').lower()
    if not any(path.endswith(ext) for ext in ALLOWED_IMAGE_EXTENSIONS):
        raise ValueError('campaign_image_url must end with .jpg, .jpeg, or .png')


def _template_payload_has_image_header(template_payload, template_record=None):
    if template_record and (template_record.header_type or '').upper() == 'IMAGE':
        return True

    if not isinstance(template_payload, dict):
        return False

    for component in template_payload.get('components', []) or []:
        if not isinstance(component, dict):
            continue
        component_type = (component.get('type') or '').upper()
        component_format = (component.get('format') or '').upper()
        if component_type == 'HEADER' and component_format == 'IMAGE':
            return True

    return False


def _inject_image_header_parameter(template_payload, image_obj):
    components = template_payload.setdefault('components', [])
    if not isinstance(components, list):
        raise ValueError('template.components must be an array')

    header_component = None
    for component in components:
        if isinstance(component, dict) and (component.get('type') or '').upper() == 'HEADER':
            header_component = component
            break

    parameter = {
        'type': 'image',
        'image': image_obj
    }

    if header_component is None:
        components.append({
            'type': 'header',
            'parameters': [parameter],
        })
        return

    header_component['parameters'] = [parameter]


def _extract_image_link_from_template_payload(template_payload):
    if not isinstance(template_payload, dict):
        return None

    components = template_payload.get('components', []) or []
    for component in components:
        if not isinstance(component, dict):
            continue
        if (component.get('type') or '').upper() != 'HEADER':
            continue

        for parameter in component.get('parameters', []) or []:
            if not isinstance(parameter, dict):
                continue
            if (parameter.get('type') or '').lower() != 'image':
                continue

            image_payload = parameter.get('image')
            if isinstance(image_payload, dict) and image_payload.get('link'):
                return image_payload.get('link')

    return None


def _prepare_template_payload_for_send(
    template_payload,
    template_record,
    waba_account,
    campaign_image_file=None,
    campaign_image_url=None,
):
    """Resolve the final template payload, handling IMAGE header assets.

    If an image file is provided, upload it to object storage and inject the
    returned public link into the template header parameters.
    """
    if not isinstance(template_payload, dict):
        raise ValueError('template payload must be an object')

    resolved_payload = copy.deepcopy(template_payload)

    requires_image_header = _template_payload_has_image_header(
        resolved_payload,
        template_record=template_record,
    )
    if not requires_image_header:
        return resolved_payload

    if campaign_image_file and campaign_image_url:
        raise ValueError('Provide either campaign_image file or campaign_image_url, not both')

    if not campaign_image_file and not campaign_image_url:
        raise ValueError('IMAGE header templates require campaign_image upload or campaign_image_url')

    image_obj = None

    if campaign_image_url:
        _validate_campaign_image_url(campaign_image_url)
        image_obj = {'link': campaign_image_url}
    else:
        _validate_campaign_image_file(campaign_image_file)
        try:
            campaign_image_link = upload_file(
                campaign_image_file,
                user_id=waba_account.user_id,
                subfolder='campaign-images'
            )
            logger.info('campaign image uploaded to object storage: %s', campaign_image_link)
            image_obj = {'link': campaign_image_link}
        finally:
            try:
                campaign_image_file.close()
            except Exception:
                pass

    _inject_image_header_parameter(resolved_payload, image_obj)
    return resolved_payload


# ---------------------------------------------------------------------------
# Conversation window helpers
# ---------------------------------------------------------------------------

def _calculate_window_expires_at(start_time):
    return start_time + timedelta(hours=CONVERSATION_WINDOW_HOURS)


def _get_window_expires_at(conversation):
    if conversation is None:
        return None

    if conversation.created_at:
        return _calculate_window_expires_at(conversation.created_at)

    return conversation.session_expires_at


def _is_conversation_expired(conversation, now=None):
    now = now or ist_now()
    window_expires_at = _get_window_expires_at(conversation)
    return window_expires_at is not None and window_expires_at <= now


def _ensure_conversation_window(conversation):
    """Ensure expiration fields are present and close expired conversations."""
    if conversation is None:
        return None

    canonical_window_expires_at = _get_window_expires_at(conversation)
    if canonical_window_expires_at and conversation.session_expires_at != canonical_window_expires_at:
        conversation.session_expires_at = canonical_window_expires_at

    now = ist_now()
    if _is_conversation_expired(conversation, now=now):
        conversation.status = 'closed'

    return conversation


def _reconcile_contact_conversations(contact_id, waba_account_id):
    """Close expired open conversations and normalize missing expiry fields."""
    conversations = (
        Conversation.query
        .filter_by(contact_id=contact_id, waba_account_id=waba_account_id)
        .order_by(desc(Conversation.created_at))
        .all()
    )

    changed = False
    now = ist_now()
    active_assigned = False
    for conversation in conversations:
        canonical_window_expires_at = _get_window_expires_at(conversation)
        if canonical_window_expires_at and conversation.session_expires_at != canonical_window_expires_at:
            conversation.session_expires_at = canonical_window_expires_at
            changed = True

        is_expired = _is_conversation_expired(conversation, now=now)
        desired_status = 'closed'
        if not is_expired and not active_assigned:
            desired_status = 'open'
            active_assigned = True

        if conversation.status != desired_status:
            conversation.status = desired_status
            changed = True

    return conversations, changed


def _is_template_message(message_type):
    return (message_type or '').lower() == 'template'


def _is_normal_message(message_type):
    normalized = (message_type or 'text').lower()
    return normalized in ('text', 'normal')


def _extract_text_from_parameter(parameter):
    """Extract a displayable text value from a template parameter object."""
    if not isinstance(parameter, dict):
        return None

    if parameter.get('text'):
        return parameter.get('text')

    if parameter.get('payload'):
        return parameter.get('payload')

    if isinstance(parameter.get('currency'), dict):
        return parameter.get('currency', {}).get('fallback_value')

    if isinstance(parameter.get('date_time'), dict):
        return parameter.get('date_time', {}).get('fallback_value')

    return None


def _format_template_message_for_storage(template_payload, template_record=None):
    """Build a UI-friendly stored body for template messages with tagged sections."""
    if not isinstance(template_payload, dict):
        return None

    template_name = template_payload.get('name')
    header_text = None
    body_text = None
    footer_text = None
    tag_values = []

    for component in template_payload.get('components', []) or []:
        if not isinstance(component, dict):
            continue

        component_type = (component.get('type') or '').upper()
        component_text = component.get('text')
        if component_type == 'HEADER' and component_text:
            header_text = component_text
        elif component_type == 'BODY' and component_text:
            body_text = component_text
        elif component_type == 'FOOTER' and component_text:
            footer_text = component_text

        parameters = component.get('parameters', []) or []
        for index, parameter in enumerate(parameters, start=1):
            param_value = _extract_text_from_parameter(parameter)
            if not param_value:
                continue
            tag_values.append(f"{component_type or 'PARAM'}_{index}:{param_value}")

    if template_record:
        header_text = header_text or template_record.header_content
        body_text = body_text or template_record.body_text
        footer_text = footer_text or template_record.footer_text

    lines = []
    if template_name:
        lines.append(f"[TEMPLATE] {template_name}")
    if header_text:
        lines.append(f"[HEADER] {header_text}")
    if body_text:
        lines.append(f"[BODY] {body_text}")
    if footer_text:
        lines.append(f"[FOOTER] {footer_text}")
    if tag_values:
        lines.append(f"[TAGS] {' | '.join(tag_values)}")

    return '\n'.join(lines) if lines else template_name


def _create_conversation(contact_id, waba_account_id, start_time=None):
    start_time = start_time or ist_now()
    conversation = Conversation(
        contact_id=contact_id,
        waba_account_id=waba_account_id,
        status='open',
        created_at=start_time,
        updated_at=start_time,
        session_expires_at=_calculate_window_expires_at(start_time)
    )
    db.session.add(conversation)
    db.session.flush()
    return conversation


def _get_latest_conversation(contact_id, waba_account_id):
    conversations, _ = _reconcile_contact_conversations(contact_id, waba_account_id)
    return conversations[0] if conversations else None


def _get_active_conversation(contact_id, waba_account_id):
    conversations, _ = _reconcile_contact_conversations(contact_id, waba_account_id)
    conversation = next((c for c in conversations if c.status == 'open'), None)
    return conversation


def _get_or_create_conversation_for_send(contact_id, waba_account_id, is_template):
    """Apply 24-hour messaging policy and return conversation or policy error payload."""
    active_conversation = _get_active_conversation(contact_id, waba_account_id)
    latest_conversation = _get_latest_conversation(contact_id, waba_account_id)

    if active_conversation:
        return active_conversation, False, None

    if not is_template:
        return None, True, {
            'error': 'CONVERSATION_WINDOW_EXPIRED',
            'message': 'The 24-hour conversation window has expired. Only template messages can be sent.',
            'TEMPLATE_ONLY': True
        }

    if latest_conversation and latest_conversation.status != 'closed':
        latest_conversation.status = 'closed'

    new_conversation = _create_conversation(contact_id, waba_account_id)
    return new_conversation, False, None


def _select_conversation_for_metadata(contact_id, waba_account_id):
    active_conversation = _get_active_conversation(contact_id, waba_account_id)
    if active_conversation:
        return active_conversation
    return _get_latest_conversation(contact_id, waba_account_id)


def _get_user_waba_accounts(user_id):
    """Fetch all WABA accounts for the current user."""
    return WabaAccount.query.filter_by(user_id=user_id).all()


def _get_last_message_for_contact(contact_id):
    """Fetch the most recent message in any conversation with this contact."""
    last_message = (
        Message.query
        .filter_by(contact_id=contact_id)
        .order_by(desc(Message.sent_at))
        .first()
    )
    return last_message


def _normalize_header_name(value):
    return ''.join(ch for ch in (value or '').strip().lower() if ch.isalnum())


def _resolve_header_index(headers, aliases):
    normalized_aliases = {_normalize_header_name(alias) for alias in aliases}
    for index, header in enumerate(headers):
        if _normalize_header_name(str(header)) in normalized_aliases:
            return index
    return None


def _normalize_phone_number(raw_value, default_country_code='91'):
    normalized_raw = raw_value
    if isinstance(normalized_raw, float) and normalized_raw.is_integer():
        normalized_raw = int(normalized_raw)

    raw_str = str(normalized_raw or '').strip()
    if raw_str.endswith('.0') and raw_str.replace('.', '', 1).isdigit():
        raw_str = raw_str[:-2]

    digits_only = ''.join(ch for ch in raw_str if ch.isdigit())
    if not digits_only:
        return None

    if digits_only.startswith('0') and len(digits_only) == 11:
        digits_only = digits_only[1:]

    if len(digits_only) == 10:
        return f"{default_country_code}{digits_only}"

    if len(digits_only) < 11 or len(digits_only) > 15:
        return None

    return digits_only


def _parse_contact_rows_from_csv(file_bytes):
    decoded_content = file_bytes.decode('utf-8-sig', errors='ignore')
    reader = csv.reader(StringIO(decoded_content))
    rows = [row for row in reader if row]
    if not rows:
        return [], False

    headers = [str(item).strip() for item in rows[0]]
    phone_idx = _resolve_header_index(headers, {'phone_number', 'phoneNumber', 'phone', 'mobile', 'mobile_number'})
    name_idx = _resolve_header_index(headers, {'name', 'full_name', 'fullname', 'contact_name'})
    has_name_column = name_idx is not None

    if phone_idx is None:
        raise ValueError('Missing phone column. Expected one of: phone_number, phoneNumber, phone')

    parsed_rows = []
    for row in rows[1:]:
        phone_value = row[phone_idx] if phone_idx < len(row) else None
        name_value = row[name_idx] if has_name_column and name_idx < len(row) else None
        parsed_rows.append({'phone': phone_value, 'name': name_value})

    return parsed_rows, has_name_column


def _parse_contact_rows_from_xlsx(file_bytes):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError('XLSX upload requires openpyxl to be installed') from exc

    workbook = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    worksheet = workbook.active

    iter_rows = worksheet.iter_rows(values_only=True)
    header_row = next(iter_rows, None)
    if not header_row:
        return [], False

    headers = [str(item).strip() if item is not None else '' for item in header_row]
    phone_idx = _resolve_header_index(headers, {'phone_number', 'phoneNumber', 'phone', 'mobile', 'mobile_number'})
    name_idx = _resolve_header_index(headers, {'name', 'full_name', 'fullname', 'contact_name'})
    has_name_column = name_idx is not None

    if phone_idx is None:
        raise ValueError('Missing phone column. Expected one of: phone_number, phoneNumber, phone')

    parsed_rows = []
    for row in iter_rows:
        row_values = list(row or [])
        phone_value = row_values[phone_idx] if phone_idx < len(row_values) else None
        name_value = row_values[name_idx] if has_name_column and name_idx < len(row_values) else None
        parsed_rows.append({'phone': phone_value, 'name': name_value})

    return parsed_rows, has_name_column


def _parse_bulk_contact_file(file_storage):
    if not file_storage or not getattr(file_storage, 'filename', None):
        raise ValueError('Missing file upload')

    filename = file_storage.filename.strip().lower()
    file_bytes = file_storage.read()
    if not file_bytes:
        raise ValueError('Uploaded file is empty')

    if filename.endswith('.csv'):
        return _parse_contact_rows_from_csv(file_bytes)
    if filename.endswith('.xlsx'):
        return _parse_contact_rows_from_xlsx(file_bytes)

    raise ValueError('Unsupported file type. Upload a .csv or .xlsx file')


def _contact_to_dict(contact, last_message=None):
    """Convert a Contact model to a dictionary with optional last message."""
    data = {
        'id': contact.id,
        'phone_number': contact.phone_number,
        'name': contact.name,
        'created_at': contact.created_at.isoformat() if contact.created_at else None,
    }

    if last_message:
        data['last_message'] = last_message.body or f"[{last_message.type}]"
        data['last_message_time'] = last_message.sent_at.isoformat() if last_message.sent_at else None
        data['last_message_type'] = last_message.type
    else:
        data['last_message'] = None
        data['last_message_time'] = None
        data['last_message_type'] = None

    return data


def _group_to_dict(group, last_message=None):
    """Convert a Group model to a dictionary with optional last message."""
    data = {
        'id': group.id,
        'name': group.name,
        'description': group.description,
        'created_at': group.created_at.isoformat() if group.created_at else None,
    }

    if last_message:
        data['last_message'] = last_message.body or f"[{last_message.type}]"
        data['last_message_time'] = last_message.sent_at.isoformat() if last_message.sent_at else None
    else:
        data['last_message'] = None
        data['last_message_time'] = None

    return data


def _message_to_dict(message):
    """Convert a Message model to a dictionary."""
    contact = Contact.query.get(message.contact_id)
    return {
        'id': message.id,
        'wamid': message.wamid,
        'body': message.body,
        'media_url': message.media_url,
        'template': (
            {'header_image_url': message.media_url}
            if message.type == 'template' and message.media_url
            else None
        ),
        'direction': message.direction,
        'type': message.type,
        'status': message.status,
        'sent_at': message.sent_at.isoformat() if message.sent_at else None,
        'delivered_at': message.delivered_at.isoformat() if message.delivered_at else None,
        'read_at': message.read_at.isoformat() if message.read_at else None,
        'created_at': ist_now().isoformat(),
        'sender_name': contact.name if contact else 'Unknown',
        'contact_id': message.contact_id,
    }


def _is_contact_blocked(contact):
    """Treat explicit opt-out contacts as blocked recipients."""
    return contact.opt_in_at is not None and not bool(contact.opt_in_status)


def _resolve_group_recipients(group):
    """Resolve and deduplicate group members, dropping invalid and blocked contacts."""
    memberships = GroupContact.query.filter_by(group_id=group.id).all()
    unique_contact_ids = []
    for membership in memberships:
        if membership.contact_id not in unique_contact_ids:
            unique_contact_ids.append(membership.contact_id)

    if not unique_contact_ids:
        return []

    contacts = Contact.query.filter(
        Contact.id.in_(unique_contact_ids),
        Contact.waba_account_id == group.waba_account_id
    ).all()
    contacts_by_id = {contact.id: contact for contact in contacts}

    recipients = []
    for contact_id in unique_contact_ids:
        contact = contacts_by_id.get(contact_id)
        if not contact:
            continue
        if not contact.phone_number:
            continue
        if _is_contact_blocked(contact):
            continue
        recipients.append(contact)

    return recipients


def _summarize_group_recipient_statuses(recipient_records):
    status_summary = {}
    for recipient in recipient_records:
        status_key = recipient.status or 'queued'
        status_summary[status_key] = status_summary.get(status_key, 0) + 1
    return status_summary


def _group_message_to_dict(group_message, recipient_records, sender_name=None):
    status_summary = _summarize_group_recipient_statuses(recipient_records)
    failed_count = status_summary.get('failed', 0)
    recipient_count = len(recipient_records)
    accepted_count = recipient_count - failed_count

    return {
        'id': group_message.id,
        'chat_type': 'group',
        'group_id': group_message.group_id,
        'direction': 'outbound',
        'type': group_message.message_type,
        'body': group_message.body,
        'template': group_message.template_payload,
        'status': 'failed' if failed_count == recipient_count and recipient_count else 'sent',
        'status_summary': status_summary,
        'recipient_count': recipient_count,
        'accepted_count': accepted_count,
        'failed_count': failed_count,
        'created_at': group_message.created_at.isoformat() if group_message.created_at else None,
        'sender_name': sender_name or 'Unknown',
    }


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@chat_bp.route('/contacts', methods=['GET'])
@jwt_required()
def get_contacts():
    """Fetch the initial list of tabbed contacts."""
    try:
        current_user_id = get_jwt_identity()

        waba_accounts = _get_user_waba_accounts(current_user_id)
        waba_account_ids = [w.id for w in waba_accounts]

        if not waba_account_ids:
            return success_response([], 'No contacts - configure WhatsApp account first')

        contacts = Contact.query.filter(
            Contact.waba_account_id.in_(waba_account_ids)
        ).order_by(desc(Contact.created_at)).all()

        result = []
        for contact in contacts:
            last_message = _get_last_message_for_contact(contact.id)
            result.append(_contact_to_dict(contact, last_message))

        return success_response(result, 'Contacts fetched successfully')

    except Exception as e:
        logger.error(f"Error fetching contacts: {str(e)}", exc_info=True)
        return error_response('Failed to fetch contacts', 500)


@chat_bp.route('/messages', methods=['GET'])
@jwt_required()
def get_messages():
    """Fetch paginated messages for either a contact or group chat context."""
    try:
        current_user_id = get_jwt_identity()
        contact_id = request.args.get('contact_id', type=int) or request.args.get('chatId', type=int)
        group_id = request.args.get('group_id', type=int)
        conversation_id = request.args.get('conversation_id', type=int) or request.args.get('conversationId', type=int)
        page = request.args.get('page', 1, type=int)
        limit = request.args.get('limit', 20, type=int)

        waba_accounts = _get_user_waba_accounts(current_user_id)
        waba_account_ids = [w.id for w in waba_accounts]

        if contact_id and group_id:
            return error_response('Provide either contact_id or group_id, not both', 422)
        if not contact_id and not group_id:
            return error_response('Missing contact_id or group_id', 422)

        if group_id:
            group = Group.query.filter_by(id=group_id).first()
            if not group:
                return error_response('Group not found', 404)
            if group.waba_account_id not in waba_account_ids:
                return error_response('Unauthorized access to group', 403)

            pagination = GroupMessage.query.filter_by(
                group_id=group_id,
                waba_account_id=group.waba_account_id
            ).order_by(desc(GroupMessage.created_at)).paginate(
                page=page,
                per_page=limit,
                error_out=False
            )

            group_message_ids = [message.id for message in pagination.items]
            recipient_records = GroupMessageRecipient.query.filter(
                GroupMessageRecipient.group_message_id.in_(group_message_ids)
            ).all() if group_message_ids else []

            recipients_by_message_id = defaultdict(list)
            for recipient in recipient_records:
                recipients_by_message_id[recipient.group_message_id].append(recipient)

            sender_ids = list({message.created_by for message in pagination.items if message.created_by})
            users = User.query.filter(User.id.in_(sender_ids)).all() if sender_ids else []
            users_by_id = {user.id: user for user in users}

            messages = []
            for group_message in reversed(pagination.items):
                sender = users_by_id.get(group_message.created_by)
                sender_name = sender.username if sender else None
                messages.append(_group_message_to_dict(
                    group_message,
                    recipients_by_message_id.get(group_message.id, []),
                    sender_name=sender_name
                ))

            return success_response({
                'chat_type': 'group',
                'group_id': group_id,
                'messages': messages,
                'TEMPLATE_ONLY': False,
                'page': page,
                'limit': limit,
                'total': pagination.total,
                'has_more': page < pagination.pages
            }, 'Messages fetched successfully')

        contact = Contact.query.filter_by(id=contact_id).first()
        if not contact:
            return error_response('Contact not found', 404)
        if contact.waba_account_id not in waba_account_ids:
            return error_response('Unauthorized access to contact', 403)

        conversations, changed = _reconcile_contact_conversations(contact_id, contact.waba_account_id)
        if changed:
            db.session.commit()

        if conversation_id:
            conversation = Conversation.query.filter_by(
                id=conversation_id,
                contact_id=contact_id,
                waba_account_id=contact.waba_account_id
            ).first()
            conversation_ids = [conversation.id] if conversation else []
        else:
            conversation = _select_conversation_for_metadata(contact_id, contact.waba_account_id)
            conversation_ids = [c.id for c in conversations]

        if not conversation or not conversation_ids:
            return success_response({
                'chat_type': 'contact',
                'contact_id': contact_id,
                'conversation_id': None,
                'conversation_start_time': None,
                'window_expires_at': None,
                'TEMPLATE_ONLY': True,
                'messages': [],
                'page': page,
                'limit': limit,
                'total': 0,
                'has_more': False
            }, 'No messages for this contact')

        pagination = Message.query.filter(
            Message.conversation_id.in_(conversation_ids)
        ).order_by(desc(Message.sent_at)).paginate(
            page=page,
            per_page=limit,
            error_out=False
        )

        messages = [_message_to_dict(msg) for msg in reversed(pagination.items)]

        window_expires_at = _get_window_expires_at(conversation)
        template_only = _is_conversation_expired(conversation)

        return success_response({
            'chat_type': 'contact',
            'contact_id': contact_id,
            'conversation_id': conversation.id,
            'conversation_start_time': conversation.created_at.isoformat() if conversation.created_at else None,
            'window_expires_at': window_expires_at.isoformat() if window_expires_at else None,
            'TEMPLATE_ONLY': template_only,
            'messages': messages,
            'page': page,
            'limit': limit,
            'total': pagination.total,
            'has_more': page < pagination.pages
        }, 'Messages fetched successfully')

    except Exception as e:
        logger.error(f"Error fetching messages: {str(e)}", exc_info=True)
        return error_response('Failed to fetch messages', 500)


@chat_bp.route('/messages', methods=['POST'])
@chat_bp.route('/messages/send', methods=['POST'])
@jwt_required()
def send_message():
    """Send a message to either a contact chat or a group chat context.

    Accepts both:
      - application/json  (no image upload; use campaign_image_url for image templates)
      - multipart/form-data  (supports header_image / file upload for image templates)
    """
    try:
        logger.info("Received request to send message")
        current_user_id = get_jwt_identity()
        data, uploaded_campaign_image = _parse_send_message_request_payload()

        try:
            contact_id = _coerce_int(data.get('contact_id'), 'contact_id')
            group_id = _coerce_int(data.get('group_id'), 'group_id')
        except ValueError as parse_error:
            return error_response(str(parse_error), 422)

        chat_type = (data.get('chat_type') or ('group' if data.get('group_id') else 'contact')).lower()
        message_type = (data.get('type') or 'text').lower()
        body = data.get('body')
        template_payload = data.get('template')
        campaign_image_url = (data.get('campaign_image_url') or '').strip() or None
        is_template = _is_template_message(message_type)

        if chat_type not in ('contact', 'group'):
            return error_response('chat_type must be either contact or group', 422)

        if chat_type == 'group' and not group_id:
            return error_response('group_id is required when chat_type is group', 422)

        if chat_type == 'contact' and not contact_id:
            return error_response('contact_id is required when chat_type is contact', 422)

        if is_template and not template_payload:
            return error_response('Missing template payload for template message', 422)

        if _is_normal_message(message_type) and not body:
            return error_response('Missing body for normal message', 422)

        if not is_template and not _is_normal_message(message_type):
            return error_response('Unsupported message type', 422)

        waba_accounts = _get_user_waba_accounts(current_user_id)
        waba_account_ids = [w.id for w in waba_accounts]
        if not waba_account_ids:
            return error_response('No WhatsApp account configured', 400)

        if chat_type == 'contact':
            contact = Contact.query.filter_by(id=contact_id).first()
            if not contact:
                return error_response('Contact not found', 404)
            if contact.waba_account_id not in waba_account_ids:
                return error_response('Unauthorized access to contact', 403)

            waba_account = WabaAccount.query.filter_by(id=contact.waba_account_id).first()
            if not waba_account:
                return error_response('WABA account not found for this contact', 404)

            _, timers_changed = _reconcile_contact_conversations(contact_id, contact.waba_account_id)

            conversation, _, policy_error = _get_or_create_conversation_for_send(
                contact_id,
                contact.waba_account_id,
                is_template=is_template
            )

            if policy_error:
                if timers_changed:
                    db.session.commit()
                return policy_error, 403

            if is_template:
                template_record = None
                if isinstance(template_payload, dict) and template_payload.get('name'):
                    template_record = (
                        Template.query
                        .filter_by(
                            waba_account_id=contact.waba_account_id,
                            template_name=template_payload.get('name')
                        )
                        .order_by(desc(Template.id))
                        .first()
                    )

                # Upload campaign image to object storage (if provided) and inject
                # the resulting public link into template header parameters.
                resolved_template_payload = _prepare_template_payload_for_send(
                    template_payload,
                    template_record,
                    waba_account,
                    campaign_image_file=uploaded_campaign_image,
                    campaign_image_url=campaign_image_url,
                )
                logger.info('resolved_template_payload: %s', resolved_template_payload)

                formatted_template_body = _format_template_message_for_storage(
                    resolved_template_payload,
                    template_record=template_record
                )

                wa_payload = {
                    'messaging_product': 'whatsapp',
                    'to': contact.phone_number,
                    'type': 'template',
                    'template': resolved_template_payload
                }
            else:
                formatted_template_body = None
                wa_payload = {
                    'messaging_product': 'whatsapp',
                    'to': contact.phone_number,
                    'type': 'text',
                    'text': {
                        'body': body
                    }
                }

            wa_message_id, wa_status_code, wa_error = send_wa_text_message(
                waba_account.phone_number_id,
                waba_account.access_token,
                wa_payload,
                is_template=is_template
            )

            if wa_error:
                logger.error(f"Unable to send WA message. status_code={wa_status_code}, error={wa_error}")
                return error_response('Unable to send WA message', wa_status_code or 502)

            message = Message(
                waba_account_id=contact.waba_account_id,
                conversation_id=conversation.id,
                contact_id=contact_id,
                wamid=wa_message_id,
                direction='outbound',
                type='template' if is_template else 'text',
                body=body if not is_template else formatted_template_body,
                media_url=(
                    _extract_image_link_from_template_payload(resolved_template_payload)
                    if is_template else None
                ),
                status='sent',
                sent_at=ist_now(),
            )

            db.session.add(message)
            db.session.commit()

            status_summary = {'sent': 1}
            return success_response({
                'message_id': message.id,
                'chat_type': 'contact',
                'contact_id': contact_id,
                'recipient_count': 1,
                'accepted_count': 1,
                'failed_count': 0,
                'status_summary': status_summary,
                'message': _message_to_dict(message),
            }, 'Message sent successfully', 201)

        # --- group path ---
        group = Group.query.filter_by(id=group_id).first()
        if not group:
            return error_response('Group not found', 404)
        if group.waba_account_id not in waba_account_ids:
            return error_response('Unauthorized access to group', 403)

        waba_account = WabaAccount.query.filter_by(id=group.waba_account_id).first()
        if not waba_account:
            return error_response('WABA account not found for this group', 404)

        recipients = _resolve_group_recipients(group)
        if not recipients:
            return error_response('This group has no members. Add contacts to the group first.', 400)

        template_record = None
        if is_template and isinstance(template_payload, dict) and template_payload.get('name'):
            template_record = (
                Template.query
                .filter_by(
                    waba_account_id=group.waba_account_id,
                    template_name=template_payload.get('name')
                )
                .order_by(desc(Template.id))
                .first()
            )

        # For group sends with an image template: upload the image ONCE up front,
        # then reuse the returned object-storage link for every recipient (avoids
        # re-uploading the same file N times). We do this by resolving the template
        # payload here before entering the per-recipient loop.
        resolved_group_template_payload = None
        if is_template:
            resolved_group_template_payload = _prepare_template_payload_for_send(
                template_payload,
                template_record,
                waba_account,
                campaign_image_file=uploaded_campaign_image,
                campaign_image_url=campaign_image_url,
            )

        formatted_template_body = _format_template_message_for_storage(
            resolved_group_template_payload,
            template_record=template_record
        ) if is_template else None

        group_message = GroupMessage(
            waba_account_id=group.waba_account_id,
            group_id=group.id,
            message_type='template' if is_template else 'text',
            body=formatted_template_body if is_template else body,
            template_payload=resolved_group_template_payload if is_template else None,
            created_by=current_user_id,
            created_at=ist_now(),
        )
        db.session.add(group_message)
        db.session.flush()

        recipient_records = []
        for recipient_contact in recipients:
            recipient_record = GroupMessageRecipient(
                waba_account_id=group.waba_account_id,
                group_message_id=group_message.id,
                contact_id=recipient_contact.id,
                status='queued',
                queued_at=ist_now(),
            )
            db.session.add(recipient_record)
            recipient_records.append(recipient_record)

            conversation, _, policy_error = _get_or_create_conversation_for_send(
                recipient_contact.id,
                group.waba_account_id,
                is_template=is_template
            )
            if policy_error:
                recipient_record.status = 'failed'
                recipient_record.error_code = policy_error.get('error')
                recipient_record.error_text = policy_error.get('message')
                recipient_record.failed_at = ist_now()
                continue

            if is_template:
                wa_payload = {
                    'messaging_product': 'whatsapp',
                    'to': recipient_contact.phone_number,
                    'type': 'template',
                    'template': resolved_group_template_payload
                }
            else:
                wa_payload = {
                    'messaging_product': 'whatsapp',
                    'to': recipient_contact.phone_number,
                    'type': 'text',
                    'text': {
                        'body': body
                    }
                }

            wa_message_id, wa_status_code, wa_error = send_wa_text_message(
                waba_account.phone_number_id,
                waba_account.access_token,
                wa_payload,
                is_template=is_template
            )

            if wa_error:
                recipient_record.status = 'failed'
                recipient_record.error_code = str(wa_status_code or 'WA_SEND_FAILED')
                recipient_record.error_text = wa_error
                recipient_record.failed_at = ist_now()
                continue

            recipient_record.provider_message_id = wa_message_id
            recipient_record.status = 'sent'
            recipient_record.sent_at = ist_now()

            db.session.add(Message(
                waba_account_id=group.waba_account_id,
                conversation_id=conversation.id,
                contact_id=recipient_contact.id,
                wamid=wa_message_id,
                direction='outbound',
                type='template' if is_template else 'text',
                body=formatted_template_body if is_template else body,
                media_url=(
                    _extract_image_link_from_template_payload(resolved_group_template_payload)
                    if is_template else None
                ),
                status='sent',
                sent_at=ist_now(),
            ))

        db.session.commit()

        status_summary = _summarize_group_recipient_statuses(recipient_records)
        failed_count = status_summary.get('failed', 0)
        recipient_count = len(recipient_records)
        accepted_count = recipient_count - failed_count

        return success_response({
            'message_id': group_message.id,
            'chat_type': 'group',
            'group_id': group.id,
            'recipient_count': recipient_count,
            'accepted_count': accepted_count,
            'failed_count': failed_count,
            'status_summary': status_summary,
        }, 'Group message send processed', 201)

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error sending message: {str(e)}", exc_info=True)
        return error_response('Failed to send message', 500)


@chat_bp.route('/contacts/add', methods=['POST'])
@jwt_required()
def add_contact():
    """Save a newly created contact."""
    try:
        current_user_id = get_jwt_identity()
        data = request.get_json()

        phone_number = data.get('phone', '').strip()
        name = data.get('name', '').strip() or None
        waba_id = data.get('waba_id')

        if not phone_number:
            return error_response('Missing phone_number', 400)

        if not phone_number.isdigit() or len(phone_number) != 10:
            return error_response('Phone number must be exactly 10 digits', 400)

        phone_number = '91' + phone_number

        existing_contact = Contact.query.filter_by(phone_number=phone_number).first()
        if existing_contact:
            return error_response('Contact with this phone number already exists', 409)

        if waba_id:
            waba_account = WabaAccount.query.filter_by(waba_id=waba_id, user_id=current_user_id).first()
        else:
            waba_account = WabaAccount.query.filter_by(user_id=current_user_id).first()

        if not waba_account:
            return error_response('No WhatsApp account configured', 400)

        contact = Contact(
            phone_number=phone_number,
            name=name,
            waba_account_id=waba_account.id,
        )

        db.session.add(contact)
        db.session.commit()

        logger.info(f"New contact added: {phone_number}")

        return success_response(_contact_to_dict(contact), 'Contact added successfully', 201)

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error adding contact: {str(e)}", exc_info=True)
        return error_response('Failed to add contact', 500)


@chat_bp.route('/contacts/bulk', methods=['POST'])
@jwt_required()
def bulk_add_contacts():
    """Bulk import contacts from CSV/XLSX multipart upload."""
    try:
        current_user_id = get_jwt_identity()
        waba_id = (request.form.get('waba_id') or '').strip() or None
        uploaded_file = request.files.get('file')

        if waba_id:
            waba_account = WabaAccount.query.filter_by(waba_id=waba_id, user_id=current_user_id).first()
        else:
            waba_account = WabaAccount.query.filter_by(user_id=current_user_id).first()

        if not waba_account:
            return error_response('No WhatsApp account configured', 400)

        parsed_rows, has_name_column = _parse_bulk_contact_file(uploaded_file)
        if not parsed_rows:
            return success_response({
                'created': 0,
                'updated': 0,
                'skipped_invalid': 0,
                'skipped_duplicates_in_file': 0,
                'skipped_existing_other_account': 0,
                'processed_rows': 0,
            }, 'No rows found in uploaded file')

        normalized_phone_numbers = set()
        for row in parsed_rows:
            normalized_number = _normalize_phone_number(row.get('phone'))
            if normalized_number:
                normalized_phone_numbers.add(normalized_number)

        existing_contacts = Contact.query.filter(
            Contact.phone_number.in_(list(normalized_phone_numbers))
        ).all() if normalized_phone_numbers else []
        existing_by_phone = {contact.phone_number: contact for contact in existing_contacts}

        created_count = 0
        updated_count = 0
        skipped_invalid = 0
        skipped_duplicates_in_file = 0
        skipped_existing_other_account = 0
        seen_numbers = set()

        for row in parsed_rows:
            normalized_number = _normalize_phone_number(row.get('phone'))
            if not normalized_number:
                skipped_invalid += 1
                continue

            if normalized_number in seen_numbers:
                skipped_duplicates_in_file += 1
                continue
            seen_numbers.add(normalized_number)

            raw_name = row.get('name')
            parsed_name = str(raw_name).strip() if raw_name is not None else ''
            name_value = parsed_name or normalized_number
            existing_contact = existing_by_phone.get(normalized_number)

            if existing_contact:
                if existing_contact.waba_account_id != waba_account.id:
                    skipped_existing_other_account += 1
                    continue

                if has_name_column and parsed_name and existing_contact.name != parsed_name:
                    existing_contact.name = parsed_name
                    updated_count += 1
                continue

            db.session.add(Contact(
                phone_number=normalized_number,
                name=name_value,
                waba_account_id=waba_account.id,
            ))
            created_count += 1

        db.session.commit()

        payload = {
            'created': created_count,
            'updated': updated_count,
            'skipped_invalid': skipped_invalid,
            'skipped_duplicates_in_file': skipped_duplicates_in_file,
            'skipped_existing_other_account': skipped_existing_other_account,
            'processed_rows': len(parsed_rows),
        }
        status_code = 201 if created_count else 200
        return success_response(payload, 'Bulk contact import completed', status_code)

    except ValueError as e:
        db.session.rollback()
        return error_response(str(e), 400)
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error importing contacts in bulk: {str(e)}", exc_info=True)
        return error_response('Failed to import contacts', 500)


@chat_bp.route('/contacts/<int:contact_id>', methods=['PUT'])
@jwt_required()
def update_contact_name(contact_id):
    """Update an existing contact's display name."""
    try:
        current_user_id = get_jwt_identity()
        data = request.get_json() or {}

        if 'name' not in data:
            return error_response('Missing name', 400)

        name = (data.get('name') or '').strip()
        if not name:
            return error_response('Name cannot be empty', 400)

        contact = Contact.query.filter_by(id=contact_id).first()
        if not contact:
            return error_response('Contact not found', 404)

        waba_accounts = _get_user_waba_accounts(current_user_id)
        waba_account_ids = [w.id for w in waba_accounts]
        if contact.waba_account_id not in waba_account_ids:
            return error_response('Unauthorized access to contact', 403)

        contact.name = name
        db.session.commit()

        logger.info(f"Contact updated: id={contact_id}")
        return success_response(_contact_to_dict(contact), 'Contact updated successfully')

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error updating contact {contact_id}: {str(e)}", exc_info=True)
        return error_response('Failed to update contact', 500)


def _group_contacts_to_dict(group_id):
    """Return contacts attached to a group as lightweight dictionaries."""
    group_contacts = GroupContact.query.filter_by(group_id=group_id).all()
    contact_ids = [gc.contact_id for gc in group_contacts]
    if not contact_ids:
        return []

    contacts = Contact.query.filter(Contact.id.in_(contact_ids)).all()
    return [
        {
            'id': contact.id,
            'name': contact.name,
            'phone_number': contact.phone_number,
        }
        for contact in contacts
    ]


def _group_detail_to_dict(group):
    """Convert Group model to response payload with contacts array."""
    contacts = _group_contacts_to_dict(group.id)
    return {
        'id': group.id,
        'name': group.name,
        'description': group.description,
        'waba_account_id': group.waba_account_id,
        'created_by': group.created_by,
        'created_at': group.created_at.isoformat() if group.created_at else None,
        'contact_ids': [contact['id'] for contact in contacts],
        'contacts': contacts,
    }


def _validate_group_contacts(contact_ids, user_waba_account_ids):
    """Validate contact IDs belong to user accounts and one WABA account."""
    if not isinstance(contact_ids, list):
        return None, None, 'contact_ids must be an array'

    normalized_ids = []
    for contact_id in contact_ids:
        if not isinstance(contact_id, int):
            return None, None, 'Each contact_id must be an integer'
        if contact_id not in normalized_ids:
            normalized_ids.append(contact_id)

    if not normalized_ids:
        return [], None, None

    contacts = Contact.query.filter(Contact.id.in_(normalized_ids)).all()
    if len(contacts) != len(normalized_ids):
        return None, None, 'One or more contact_ids are invalid'

    waba_ids = {contact.waba_account_id for contact in contacts}
    if len(waba_ids) > 1:
        return None, None, 'All contacts must belong to the same WABA account'

    resolved_waba_account_id = next(iter(waba_ids)) if waba_ids else None
    if resolved_waba_account_id and resolved_waba_account_id not in user_waba_account_ids:
        return None, None, 'One or more contacts do not belong to your account'

    return normalized_ids, resolved_waba_account_id, None


@chat_bp.route('/groups', methods=['POST'])
@jwt_required()
def create_group_crud():
    """Create a group with contact_ids array membership."""
    try:
        current_user_id = get_jwt_identity()
        data = request.get_json() or {}

        group_name = (data.get('name') or '').strip()
        description = (data.get('description') or '').strip() or None
        contact_ids = data.get('contact_ids') or []

        if not group_name:
            return error_response('Missing group name', 400)

        waba_accounts = _get_user_waba_accounts(current_user_id)
        user_waba_account_ids = [w.id for w in waba_accounts]
        if not user_waba_account_ids:
            return error_response('No WhatsApp account configured', 400)

        normalized_contact_ids, resolved_waba_account_id, validation_error = _validate_group_contacts(
            contact_ids,
            user_waba_account_ids
        )
        if validation_error:
            return error_response(validation_error, 400)

        waba_account_id = resolved_waba_account_id or user_waba_account_ids[0]

        group = Group(
            name=group_name,
            description=description,
            waba_account_id=waba_account_id,
            created_by=current_user_id,
        )
        db.session.add(group)
        db.session.flush()

        for contact_id in normalized_contact_ids:
            db.session.add(GroupContact(
                waba_account_id=waba_account_id,
                group_id=group.id,
                contact_id=contact_id,
            ))

        db.session.commit()
        return success_response(_group_detail_to_dict(group), 'Group created successfully', 201)

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error creating group: {str(e)}", exc_info=True)
        return error_response('Failed to create group', 500)


@chat_bp.route('/groups', methods=['GET'])
@jwt_required()
def get_groups_crud():
    """List groups with member contact_ids and contact details."""
    try:
        current_user_id = get_jwt_identity()
        waba_accounts = _get_user_waba_accounts(current_user_id)
        waba_account_ids = [w.id for w in waba_accounts]

        if not waba_account_ids:
            return success_response([], 'No groups - configure WhatsApp account first')

        groups = Group.query.filter(
            Group.waba_account_id.in_(waba_account_ids)
        ).order_by(desc(Group.created_at)).all()

        return success_response([
            _group_detail_to_dict(group)
            for group in groups
        ], 'Groups fetched successfully')

    except Exception as e:
        logger.error(f"Error fetching groups (CRUD): {str(e)}", exc_info=True)
        return error_response('Failed to fetch groups', 500)


@chat_bp.route('/groups/<int:group_id>', methods=['GET'])
@jwt_required()
def get_group_by_id_crud(group_id):
    """Fetch one group with contacts for CRUD workflows."""
    try:
        current_user_id = get_jwt_identity()
        waba_accounts = _get_user_waba_accounts(current_user_id)
        waba_account_ids = [w.id for w in waba_accounts]

        group = Group.query.filter_by(id=group_id).first()
        if not group:
            return error_response('Group not found', 404)

        if group.waba_account_id not in waba_account_ids:
            return error_response('Unauthorized access to group', 403)

        return success_response(_group_detail_to_dict(group), 'Group fetched successfully')

    except Exception as e:
        logger.error(f"Error fetching group {group_id}: {str(e)}", exc_info=True)
        return error_response('Failed to fetch group', 500)


@chat_bp.route('/groups/<int:group_id>/name', methods=['PUT'])
@jwt_required()
def update_group_name(group_id):
    """Update an existing group's name only."""
    try:
        current_user_id = get_jwt_identity()
        data = request.get_json() or {}

        if 'name' not in data:
            return error_response('Missing name', 400)

        group_name = (data.get('name') or '').strip()
        if not group_name:
            return error_response('Group name cannot be empty', 400)

        waba_accounts = _get_user_waba_accounts(current_user_id)
        user_waba_account_ids = [w.id for w in waba_accounts]

        group = Group.query.filter_by(id=group_id).first()
        if not group:
            return error_response('Group not found', 404)

        if group.waba_account_id not in user_waba_account_ids:
            return error_response('Unauthorized access to group', 403)

        group.name = group_name
        db.session.commit()

        logger.info(f"Group updated: id={group_id}")
        return success_response(_group_detail_to_dict(group), 'Group name updated successfully')

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error updating group name {group_id}: {str(e)}", exc_info=True)
        return error_response('Failed to update group name', 500)


@chat_bp.route('/groups/<int:group_id>', methods=['PUT'])
@jwt_required()
def update_group_crud(group_id):
    """Update group profile and replace membership using contact_ids array."""
    try:
        current_user_id = get_jwt_identity()
        data = request.get_json() or {}

        waba_accounts = _get_user_waba_accounts(current_user_id)
        user_waba_account_ids = [w.id for w in waba_accounts]

        group = Group.query.filter_by(id=group_id).first()
        if not group:
            return error_response('Group not found', 404)

        if group.waba_account_id not in user_waba_account_ids:
            return error_response('Unauthorized access to group', 403)

        if 'name' in data:
            group_name = (data.get('name') or '').strip()
            if not group_name:
                return error_response('Group name cannot be empty', 400)
            group.name = group_name

        if 'description' in data:
            group.description = (data.get('description') or '').strip() or None

        if 'contact_ids' in data:
            normalized_contact_ids, resolved_waba_account_id, validation_error = _validate_group_contacts(
                data.get('contact_ids') or [],
                user_waba_account_ids
            )
            if validation_error:
                return error_response(validation_error, 400)

            if resolved_waba_account_id and resolved_waba_account_id != group.waba_account_id:
                return error_response('contact_ids belong to a different WABA account than this group', 400)

            GroupContact.query.filter_by(group_id=group.id).delete()
            for contact_id in normalized_contact_ids:
                db.session.add(GroupContact(
                    waba_account_id=group.waba_account_id,
                    group_id=group.id,
                    contact_id=contact_id,
                ))

        db.session.commit()
        return success_response(_group_detail_to_dict(group), 'Group updated successfully')

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error updating group {group_id}: {str(e)}", exc_info=True)
        return error_response('Failed to update group', 500)


@chat_bp.route('/groups/<int:group_id>', methods=['DELETE'])
@jwt_required()
def delete_group_crud(group_id):
    """Delete a group and all its memberships."""
    try:
        current_user_id = get_jwt_identity()
        waba_accounts = _get_user_waba_accounts(current_user_id)
        user_waba_account_ids = [w.id for w in waba_accounts]

        group = Group.query.filter_by(id=group_id).first()
        if not group:
            return error_response('Group not found', 404)

        if group.waba_account_id not in user_waba_account_ids:
            return error_response('Unauthorized access to group', 403)

        GroupContact.query.filter_by(group_id=group.id).delete()
        db.session.delete(group)
        db.session.commit()

        return success_response({'id': group_id}, 'Group deleted successfully')

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error deleting group {group_id}: {str(e)}", exc_info=True)
        return error_response('Failed to delete group', 500)
